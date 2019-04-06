import openpyxl

def readExceltoText(dir, excel):
    # Excelを開いてアクティブシートを取得
    book = openpyxl.load_workbook(dir + excel)
    active_sheet = book.active

    # 4～50行目までループしてセル値を取得
    outputList = []
    for index in range(4, 50):
        # 値がある場合だけ処理
        if str(active_sheet.cell(column=8, row=index).value) != "None":
            outputText = str(active_sheet.cell(column=8, row=index).value)
            outputText = outputText + " " + str(active_sheet.cell(column=2, row=index).value)
            outputText = outputText + " " + str(active_sheet.cell(column=3, row=index).value)
            outputText = outputText + " " + str(active_sheet.cell(column=4, row=index).value)
            outputText = outputText + " " + str(active_sheet.cell(column=16, row=index).value) + "日"

            #値を配列にして格納
            outputList.append(outputText)

    #配列を返す
    return outputList

readExceltoText("C:/ZZ_Health_Check/Python/Samples/", "TEST用_20190406.xlsx")
